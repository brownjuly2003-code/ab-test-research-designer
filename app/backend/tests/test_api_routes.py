from pathlib import Path
import sys
import uuid

from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import app.backend.app.main as main_module
from app.backend.app.config import get_settings
from app.backend.app.main import create_app
from app.backend.app.llm.adapter import LocalOrchestratorAdapter
from app.backend.app.repository import ProjectRepository


def _full_payload() -> dict:
    return {
        "project": {
            "project_name": "Checkout redesign",
            "domain": "e-commerce",
            "product_type": "web app",
            "platform": "web",
            "market": "US",
            "project_description": "We want to test a simplified checkout flow.",
        },
        "hypothesis": {
            "change_description": "Reduce checkout from 4 steps to 2",
            "target_audience": "new users on web",
            "business_problem": "checkout abandonment is high",
            "hypothesis_statement": "If we simplify checkout, conversion will increase.",
            "what_to_validate": "impact on conversion",
            "desired_result": "statistically meaningful uplift",
        },
        "setup": {
            "experiment_type": "ab",
            "randomization_unit": "user",
            "traffic_split": [50, 50],
            "expected_daily_traffic": 12000,
            "audience_share_in_test": 0.6,
            "variants_count": 2,
            "inclusion_criteria": "new users only",
            "exclusion_criteria": "internal staff",
        },
        "metrics": {
            "primary_metric_name": "purchase_conversion",
            "metric_type": "binary",
            "baseline_value": 0.042,
            "expected_uplift_pct": 8,
            "mde_pct": 5,
            "alpha": 0.05,
            "power": 0.8,
            "std_dev": None,
            "secondary_metrics": ["add_to_cart_rate"],
            "guardrail_metrics": [
                {
                    "name": "Payment error rate",
                    "metric_type": "binary",
                    "baseline_rate": 2.4,
                },
                {
                    "name": "Refund value",
                    "metric_type": "continuous",
                    "baseline_mean": 18.0,
                    "std_dev": 6.5,
                },
            ],
        },
        "constraints": {
            "seasonality_present": True,
            "active_campaigns_present": False,
            "returning_users_present": True,
            "interference_risk": "medium",
            "technical_constraints": "legacy event logging",
            "legal_or_ethics_constraints": "none",
            "known_risks": "tracking quality",
            "deadline_pressure": "medium",
            "long_test_possible": True,
        },
        "additional_context": {
            "llm_context": "Previous tests showed mixed results.",
        },
    }


def _continuous_full_payload() -> dict:
    payload = _full_payload()
    payload["metrics"] = {
        "primary_metric_name": "avg_order_value",
        "metric_type": "continuous",
        "baseline_value": 45.0,
        "expected_uplift_pct": 6,
        "mde_pct": 4.4444444444,
        "alpha": 0.05,
        "power": 0.8,
        "std_dev": 12.0,
        "cuped_pre_experiment_std": 12.0,
        "cuped_correlation": 0.5,
        "secondary_metrics": ["revenue_per_user"],
        "guardrail_metrics": [
            {
                "name": "Refund value",
                "metric_type": "continuous",
                "baseline_mean": 18.0,
                "std_dev": 6.5,
            }
        ],
    }
    return payload


def _saved_analysis_payload(
    *,
    metric_type: str = "binary",
    total_sample_size: int = 200,
    estimated_duration_days: int = 10,
    executive_summary: str = "Summary",
    warning_codes: list[str] | None = None,
    assumptions: list[str] | None = None,
    risk_highlights: dict[str, list[str]] | None = None,
    recommendation_highlights: dict[str, list[str]] | None = None,
) -> dict:
    sample_size_per_variant = total_sample_size // 2
    baseline_value = 0.042 if metric_type == "binary" else 45.0
    primary_metric = "purchase_conversion" if metric_type == "binary" else "avg_order_value"
    resolved_warning_codes = warning_codes if warning_codes is not None else ["SEASONALITY_PRESENT"]
    resolved_assumptions = assumptions if assumptions is not None else ["Baseline is stable"]
    resolved_risks = risk_highlights if risk_highlights is not None else {
        "statistical": ["Power tradeoff"],
        "product": ["Expected result depends on user behavior."],
        "technical": ["legacy event logging"],
        "operational": ["tracking quality"],
    }
    resolved_recommendations = recommendation_highlights if recommendation_highlights is not None else {
        "before_launch": ["Verify tracking"],
        "during_test": ["Watch SRM"],
        "after_test": ["Segment the result"],
    }

    return {
        "calculations": {
            "calculation_summary": {
                "metric_type": metric_type,
                "baseline_value": baseline_value,
                "mde_pct": 5,
                "mde_absolute": 0.0021 if metric_type == "binary" else 2.0,
                "alpha": 0.05,
                "power": 0.8,
            },
            "results": {
                "sample_size_per_variant": sample_size_per_variant,
                "total_sample_size": total_sample_size,
                "effective_daily_traffic": 5000,
                "estimated_duration_days": estimated_duration_days,
            },
            "assumptions": resolved_assumptions,
            "warnings": [
                {
                    "code": code,
                    "severity": "medium",
                    "message": f"{code} may affect the result.",
                    "source": "rules_engine",
                }
                for code in resolved_warning_codes
            ],
        },
        "report": {
            "executive_summary": executive_summary,
            "calculations": {
                "sample_size_per_variant": sample_size_per_variant,
                "total_sample_size": total_sample_size,
                "estimated_duration_days": estimated_duration_days,
                "assumptions": resolved_assumptions,
            },
            "experiment_design": {
                "variants": [
                    {"name": "Control", "description": "current"},
                    {"name": "Treatment", "description": "candidate"},
                ],
                "randomization_unit": "user",
                "traffic_split": [50, 50],
                "target_audience": "new users on web",
                "inclusion_criteria": "new users only",
                "exclusion_criteria": "internal staff",
                "recommended_duration_days": estimated_duration_days,
                "stopping_conditions": ["planned duration reached"],
            },
            "metrics_plan": {
                "primary": [primary_metric],
                "secondary": ["add_to_cart_rate"],
                "guardrail": ["payment_error_rate"],
                "diagnostic": ["assignment_rate"],
            },
            "guardrail_metrics": [
                {
                    "name": "Payment error rate",
                    "metric_type": "binary",
                    "baseline": 2.4,
                    "detectable_mde_pp": 0.321,
                    "note": "Can detect a 0.321 pp change",
                }
            ],
            "risks": resolved_risks,
            "recommendations": resolved_recommendations,
            "open_questions": ["Will mobile respond differently?"],
        },
        "advice": {
            "available": True,
            "provider": "local_orchestrator",
            "model": "offline",
            "advice": {
                "brief_assessment": "Feasible with monitoring.",
                "key_risks": ["Tracking quality"],
                "design_improvements": ["Validate assignment logging"],
                "metric_recommendations": ["Track checkout step completion"],
                "interpretation_pitfalls": ["Do not stop early"],
                "additional_checks": ["Verify exposure balance"],
            },
            "raw_text": None,
            "error": None,
            "error_code": None,
        },
    }


def _saved_observed_results(metric_type: str = "binary") -> dict:
    if metric_type == "continuous":
        return {
            "request": {
                "metric_type": "continuous",
                "continuous": {
                    "control_mean": 45.0,
                    "control_std": 12.0,
                    "control_n": 200,
                    "treatment_mean": 47.2,
                    "treatment_std": 12.3,
                    "treatment_n": 205,
                    "alpha": 0.05,
                },
            },
            "analysis": {
                "metric_type": "continuous",
                "observed_effect": 2.2,
                "observed_effect_relative": 4.89,
                "control_rate": None,
                "treatment_rate": None,
                "ci_lower": 0.4,
                "ci_upper": 4.0,
                "ci_level": 0.95,
                "p_value": 0.03,
                "test_statistic": 2.17,
                "is_significant": True,
                "power_achieved": 0.71,
                "verdict": "Ship candidate",
                "interpretation": "Treatment is above control.",
            },
            "saved_at": "2026-03-07T13:15:00Z",
        }

    return {
        "request": {
            "metric_type": "binary",
            "binary": {
                "control_conversions": 410,
                "control_users": 10000,
                "treatment_conversions": 472,
                "treatment_users": 10020,
                "alpha": 0.05,
            },
        },
        "analysis": {
            "metric_type": "binary",
            "observed_effect": 0.62,
            "observed_effect_relative": 15.12,
            "control_rate": 0.041,
            "treatment_rate": 0.0472,
            "ci_lower": 0.12,
            "ci_upper": 1.11,
            "ci_level": 0.95,
            "p_value": 0.018,
            "test_statistic": 2.36,
            "is_significant": True,
            "power_achieved": 0.83,
            "verdict": "Ship candidate",
            "interpretation": "Treatment beat control.",
        },
        "saved_at": "2026-03-07T13:15:00Z",
    }


def _create_saved_project(
    client: TestClient,
    name: str,
    *,
    metric_type: str = "binary",
    total_sample_size: int = 200,
    estimated_duration_days: int = 10,
    warning_codes: list[str] | None = None,
    assumptions: list[str] | None = None,
    risk_highlights: dict[str, list[str]] | None = None,
    recommendation_highlights: dict[str, list[str]] | None = None,
) -> dict:
    payload = _full_payload() if metric_type == "binary" else _continuous_full_payload()
    payload["project"]["project_name"] = name
    payload["additional_context"]["observed_results"] = _saved_observed_results(metric_type)
    created = client.post("/api/v1/projects", json=payload)
    assert created.status_code == 200

    analysis = client.post(
        f"/api/v1/projects/{created.json()['id']}/analysis",
        json=_saved_analysis_payload(
            metric_type=metric_type,
            total_sample_size=total_sample_size,
            estimated_duration_days=estimated_duration_days,
            executive_summary=f"{name} summary",
            warning_codes=warning_codes,
            assumptions=assumptions,
            risk_highlights=risk_highlights,
            recommendation_highlights=recommendation_highlights,
        ),
    )
    assert analysis.status_code == 200
    return created.json()


def test_calculate_endpoint_returns_deterministic_payload() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "binary",
            "baseline_value": 0.042,
            "mde_pct": 5,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 12000,
            "audience_share_in_test": 0.6,
            "traffic_split": [50, 50],
            "variants_count": 2,
        },
    )

    assert response.status_code == 200
    assert response.json()["calculation_summary"]["metric_type"] == "binary"
    assert response.json()["bonferroni_note"] is None
    assert response.headers["x-request-id"]
    assert float(response.headers["x-process-time-ms"]) >= 0


def test_calculate_endpoint_returns_cuped_fields_for_continuous_metric() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "continuous",
            "baseline_value": 45.0,
            "std_dev": 12.0,
            "mde_pct": 4.4444444444,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 10000,
            "audience_share_in_test": 1.0,
            "traffic_split": [50, 50],
            "variants_count": 2,
            "cuped_pre_experiment_std": 12.0,
            "cuped_correlation": 0.5,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["cuped_std"] == 10.3923
    assert payload["cuped_variance_reduction_pct"] == 25.0
    assert payload["cuped_sample_size_per_variant"] < payload["results"]["sample_size_per_variant"]
    assert payload["cuped_duration_days"] <= payload["results"]["estimated_duration_days"]


def test_compare_multi_two_projects() -> None:
    client = TestClient(create_app())
    first_project = _create_saved_project(
        client,
        "Checkout baseline",
        total_sample_size=220,
        estimated_duration_days=9,
        warning_codes=["SEASONALITY_PRESENT", "LOW_TRAFFIC"],
        assumptions=["Baseline is stable", "Traffic split holds"],
    )
    second_project = _create_saved_project(
        client,
        "Checkout challenger",
        total_sample_size=300,
        estimated_duration_days=12,
        warning_codes=["LOW_TRAFFIC", "LONG_DURATION"],
        assumptions=["Baseline is stable", "New checkout keeps load time"],
        recommendation_highlights={
            "before_launch": ["Verify tracking"],
            "during_test": ["Watch SRM"],
            "after_test": ["Segment by device"],
        },
    )

    response = client.post(
        "/api/v1/projects/compare",
        json={"project_ids": [first_project["id"], second_project["id"]]},
    )

    assert response.status_code == 200
    payload = response.json()
    assert [item["project_name"] for item in payload["projects"]] == [
        "Checkout baseline",
        "Checkout challenger",
    ]
    assert payload["shared_warnings"] == ["LOW_TRAFFIC"]
    assert payload["shared_assumptions"] == ["Baseline is stable"]
    assert payload["sample_size_range"] == {"min": 220, "max": 300, "median": 260.0}
    assert payload["duration_range"] == {"min": 9, "max": 12, "median": 10.5}
    assert payload["metric_types_used"] == ["binary"]
    assert payload["projects"][0]["sensitivity"]["current_power"] == 0.8
    assert payload["projects"][0]["observed_results"]["metric_type"] == "binary"
    assert any("Watch SRM" in item for item in payload["recommendation_highlights"])


def test_compare_multi_three_projects() -> None:
    client = TestClient(create_app())
    first_project = _create_saved_project(
        client,
        "Baseline",
        total_sample_size=200,
        estimated_duration_days=8,
        warning_codes=["LOW_TRAFFIC"],
        assumptions=["Baseline is stable", "Traffic split holds"],
    )
    second_project = _create_saved_project(
        client,
        "Variant A",
        total_sample_size=260,
        estimated_duration_days=10,
        warning_codes=["LOW_TRAFFIC", "LONG_DURATION"],
        assumptions=["Baseline is stable", "Tracking is stable"],
    )
    third_project = _create_saved_project(
        client,
        "Variant B",
        total_sample_size=320,
        estimated_duration_days=13,
        warning_codes=["LOW_TRAFFIC", "METRIC_DRIFT"],
        assumptions=["Baseline is stable", "Audience mix stays stable"],
    )

    response = client.post(
        "/api/v1/projects/compare",
        json={"project_ids": [first_project["id"], second_project["id"], third_project["id"]]},
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["projects"]) == 3
    assert payload["shared_warnings"] == ["LOW_TRAFFIC"]
    assert payload["shared_assumptions"] == ["Baseline is stable"]
    assert payload["unique_per_project"][third_project["id"]]["warnings"] == ["METRIC_DRIFT"]
    assert payload["unique_per_project"][third_project["id"]]["assumptions"] == ["Audience mix stays stable"]


def test_compare_multi_five_projects() -> None:
    client = TestClient(create_app())
    project_ids = [
        _create_saved_project(
            client,
            f"Project {index + 1}",
            total_sample_size=200 + (index * 40),
            estimated_duration_days=8 + index,
        )["id"]
        for index in range(5)
    ]

    response = client.post(
        "/api/v1/projects/compare",
        json={"project_ids": project_ids},
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["projects"]) == 5
    assert payload["sample_size_range"] == {"min": 200, "max": 360, "median": 280}
    assert payload["duration_range"] == {"min": 8, "max": 12, "median": 10}


def test_compare_multi_six_422() -> None:
    client = TestClient(create_app())
    project_ids = [
        _create_saved_project(client, f"Project {index + 1}")["id"]
        for index in range(6)
    ]

    response = client.post(
        "/api/v1/projects/compare",
        json={"project_ids": project_ids},
    )

    assert response.status_code == 422


def test_compare_multi_single_422() -> None:
    client = TestClient(create_app())
    project_id = _create_saved_project(client, "Only project")["id"]

    response = client.post(
        "/api/v1/projects/compare",
        json={"project_ids": [project_id]},
    )

    assert response.status_code == 422


def test_compare_multi_mixed_metric_types() -> None:
    client = TestClient(create_app())
    binary_project = _create_saved_project(client, "Binary project", metric_type="binary")
    continuous_project = _create_saved_project(
        client,
        "Continuous project",
        metric_type="continuous",
        total_sample_size=420,
        estimated_duration_days=16,
    )

    response = client.post(
        "/api/v1/projects/compare",
        json={"project_ids": [binary_project["id"], continuous_project["id"]]},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["metric_types_used"] == ["binary", "continuous"]
    assert "Mixed metric types — direct effect comparison not meaningful" in payload["shared_warnings"]
    assert payload["projects"][1]["observed_results"]["metric_type"] == "continuous"


def test_calculate_endpoint_rejects_degenerate_cuped_correlation() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "continuous",
            "baseline_value": 45.0,
            "std_dev": 12.0,
            "mde_pct": 4.4444444444,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 10000,
            "audience_share_in_test": 1.0,
            "traffic_split": [50, 50],
            "variants_count": 2,
            "cuped_pre_experiment_std": 12.0,
            "cuped_correlation": 1.0,
        },
    )

    assert response.status_code == 422
    assert "less than 1" in str(response.json()["detail"])


def test_calculate_endpoint_surfaces_bonferroni_note_for_multivariant_design() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "binary",
            "baseline_value": 0.042,
            "mde_pct": 5,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 12000,
            "audience_share_in_test": 0.6,
            "traffic_split": [34, 33, 33],
            "variants_count": 3,
        },
    )

    assert response.status_code == 200
    assert response.json()["bonferroni_note"] is not None


def test_sensitivity_binary() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/sensitivity",
        json={
            "metric_type": "binary",
            "baseline_rate": 3.5,
            "mde_values": [0.5, 1.0],
            "power_values": [0.8, 0.9],
            "daily_traffic": 10000,
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert len(data["cells"]) == 4
    assert all("sample_size_per_variant" in cell for cell in data["cells"])


def test_design_endpoint_builds_report_without_llm() -> None:
    client = TestClient(create_app())

    response = client.post("/api/v1/design", json=_full_payload())

    assert response.status_code == 200
    assert response.json()["metrics_plan"]["primary"] == ["purchase_conversion"]


def test_analyze_endpoint_returns_combined_payload(monkeypatch) -> None:
    monkeypatch.setattr(
        LocalOrchestratorAdapter,
        "request_advice",
        lambda self, payload: {
            "available": True,
            "provider": "local_orchestrator",
            "model": "Claude Sonnet 4.6",
            "advice": {
                "brief_assessment": "Combined analysis available.",
                "key_risks": ["Tracking quality"],
                "design_improvements": ["Validate event schema"],
                "metric_recommendations": ["Track checkout step completion"],
                "interpretation_pitfalls": ["Do not stop early"],
                "additional_checks": ["Verify exposure balance"],
            },
            "raw_text": "{\"brief_assessment\": \"Combined analysis available.\"}",
            "error": None,
            "error_code": None,
        },
    )
    client = TestClient(create_app())

    response = client.post("/api/v1/analyze", json=_full_payload())

    assert response.status_code == 200
    payload = response.json()
    assert payload["calculations"]["calculation_summary"]["metric_type"] == "binary"
    assert payload["report"]["metrics_plan"]["primary"] == ["purchase_conversion"]
    assert payload["advice"]["available"] is True
    assert payload["advice"]["advice"]["brief_assessment"] == "Combined analysis available."
    assert len(payload["report"]["guardrail_metrics"]) == 2
    assert payload["report"]["guardrail_metrics"][0]["name"] == "Payment error rate"
    assert "detectable_mde_pp" in payload["report"]["guardrail_metrics"][0]
    assert "detectable_mde_absolute" in payload["report"]["guardrail_metrics"][1]


def test_analyze_endpoint_propagates_cuped_metrics_from_nested_payload(monkeypatch) -> None:
    monkeypatch.setattr(
        LocalOrchestratorAdapter,
        "request_advice",
        lambda self, payload: {
            "available": False,
            "provider": "local_orchestrator",
            "model": "offline",
            "advice": None,
            "raw_text": None,
            "error": "offline",
            "error_code": "request_error",
        },
    )
    client = TestClient(create_app())

    response = client.post("/api/v1/analyze", json=_continuous_full_payload())

    assert response.status_code == 200
    payload = response.json()["calculations"]
    assert payload["cuped_std"] == 10.3923
    assert payload["cuped_variance_reduction_pct"] == 25.0
    assert payload["cuped_sample_size_per_variant"] < payload["results"]["sample_size_per_variant"]


def test_analyze_endpoint_rejects_guardrail_metric_missing_required_fields() -> None:
    client = TestClient(create_app())
    payload = _full_payload()
    payload["metrics"]["guardrail_metrics"] = [
        {
            "name": "Revenue",
            "metric_type": "continuous",
        }
    ]

    response = client.post("/api/v1/analyze", json=payload)

    assert response.status_code == 422
    assert "baseline_mean" in str(response.json()["detail"])


def test_analyze_endpoint_rejects_more_than_three_guardrail_metrics() -> None:
    client = TestClient(create_app())
    payload = _full_payload()
    payload["metrics"]["guardrail_metrics"] = [
        {"name": "Bounce rate", "metric_type": "binary", "baseline_rate": 40.0},
        {"name": "Crash-free sessions", "metric_type": "binary", "baseline_rate": 99.2},
        {"name": "Revenue", "metric_type": "continuous", "baseline_mean": 12.0, "std_dev": 4.0},
        {"name": "Latency", "metric_type": "continuous", "baseline_mean": 280.0, "std_dev": 30.0},
    ]

    response = client.post("/api/v1/analyze", json=payload)

    assert response.status_code == 422
    assert "at most 3 items" in str(response.json()["detail"])


def test_llm_advice_endpoint_returns_graceful_fallback_when_orchestrator_unavailable() -> None:
    client = TestClient(create_app())

    response = client.post("/api/v1/llm/advice", json={"project_context": {"project_name": "Checkout redesign"}})

    assert response.status_code == 200
    assert response.json()["available"] is False
    assert response.json()["error_code"] is not None


def test_calculate_endpoint_rejects_mismatched_variant_configuration() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "binary",
            "baseline_value": 0.042,
            "mde_pct": 5,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 12000,
            "audience_share_in_test": 0.6,
            "traffic_split": [50, 50],
            "variants_count": 3,
        },
    )

    assert response.status_code == 422
    assert "traffic_split length must match variants_count" in str(response.json()["detail"])
    assert response.json()["error_code"] == "validation_error"
    assert response.json()["status_code"] == 422
    assert response.json()["request_id"]
    assert response.headers["x-error-code"] == "validation_error"


def test_calculate_endpoint_allows_vite_preflight_origin() -> None:
    client = TestClient(create_app())

    response = client.options(
        "/api/v1/calculate",
        headers={
            "Origin": "http://127.0.0.1:5173",
            "Access-Control-Request-Method": "POST",
            "Access-Control-Request-Headers": "content-type",
        },
    )

    assert response.status_code == 200
    assert response.headers["access-control-allow-origin"] == "http://127.0.0.1:5173"
    assert response.headers["access-control-allow-methods"] == "GET, POST, PUT, DELETE, OPTIONS"
    assert response.headers["access-control-allow-headers"] == "Accept, Accept-Language, Content-Language, Content-Type"


def test_calculate_endpoint_rejects_too_many_variants() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "binary",
            "baseline_value": 0.042,
            "mde_pct": 5,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 12000,
            "audience_share_in_test": 0.6,
            "traffic_split": [10] * 11,
            "variants_count": 11,
        },
    )

    assert response.status_code == 422
    assert "less than or equal to 10" in str(response.json()["detail"])


def test_calculate_endpoint_rejects_zero_std_dev_for_continuous_metric() -> None:
    client = TestClient(create_app())

    response = client.post(
        "/api/v1/calculate",
        json={
            "metric_type": "continuous",
            "baseline_value": 15.0,
            "std_dev": 0,
            "mde_pct": 5,
            "alpha": 0.05,
            "power": 0.8,
            "expected_daily_traffic": 12000,
            "audience_share_in_test": 0.6,
            "traffic_split": [50, 50],
            "variants_count": 2,
        },
    )

    assert response.status_code == 422
    assert "std_dev must be positive for continuous metrics" in str(response.json()["detail"])


def test_design_endpoint_returns_internal_error_for_unexpected_exception(monkeypatch) -> None:
    def explode(*args, **kwargs):
        raise KeyError("missing")

    monkeypatch.setattr(main_module, "build_experiment_report", explode)
    client = TestClient(create_app(), raise_server_exceptions=False)

    response = client.post("/api/v1/design", json=_full_payload())

    assert response.status_code == 500
    assert response.json()["detail"] == "Internal server error"
    assert response.json()["error_code"] == "internal_error"
    assert response.json()["status_code"] == 500
    assert response.json()["request_id"]


def test_diagnostics_endpoint_returns_runtime_summary(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_ENV", "test")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    monkeypatch.setenv("AB_LOG_FORMAT", "json")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.get("/api/v1/diagnostics")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "ok"
    assert payload["environment"] == "test"
    assert payload["request_timing_headers_enabled"] is True
    assert payload["storage"]["db_path"] == str(db_path)
    assert payload["storage"]["db_parent_path"] == str(db_path.parent)
    assert payload["storage"]["db_size_bytes"] >= 0
    assert payload["storage"]["disk_free_bytes"] > 0
    assert payload["storage"]["schema_version"] == 7
    assert payload["storage"]["sqlite_user_version"] == 7
    assert payload["storage"]["journal_mode"] == "WAL"
    assert payload["storage"]["synchronous"] == "NORMAL"
    assert payload["storage"]["write_probe_ok"] is True
    assert payload["storage"]["write_probe_detail"] == "BEGIN IMMEDIATE succeeded"
    assert payload["storage"]["projects_total"] == 0
    assert payload["storage"]["archived_projects_total"] == 0
    assert payload["storage"]["project_revisions_total"] == 0
    assert payload["frontend"]["serve_frontend_dist"] is False
    assert payload["llm"]["provider"] == "local_orchestrator"
    assert payload["logging"]["format"] == "json"
    assert payload["auth"]["enabled"] is False
    assert payload["auth"]["mode"] == "open"
    assert payload["auth"]["write_enabled"] is False
    assert payload["auth"]["readonly_enabled"] is False
    assert payload["guards"]["security_headers_enabled"] is True
    assert payload["guards"]["rate_limit_enabled"] is True
    assert payload["guards"]["rate_limit_requests"] == 240
    assert payload["guards"]["auth_failure_limit"] == 20
    assert payload["guards"]["max_request_body_bytes"] == 1_048_576
    assert payload["guards"]["max_workspace_body_bytes"] == 8_388_608
    assert payload["runtime"]["total_requests"] >= 0
    assert payload["runtime"]["success_responses"] >= 0
    assert payload["runtime"]["auth_rejections"] == 0
    assert payload["runtime"]["rate_limited_responses"] == 0
    assert payload["runtime"]["request_body_rejections"] == 0
    assert response.headers["x-request-id"]
    assert float(response.headers["x-process-time-ms"]) >= 0
    get_settings.cache_clear()


def test_responses_include_security_headers() -> None:
    client = TestClient(create_app())

    response = client.get("/health")

    assert response.status_code == 200
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["x-frame-options"] == "DENY"
    assert response.headers["referrer-policy"] == "no-referrer"
    assert response.headers["permissions-policy"] == "camera=(), geolocation=(), microphone=()"
    assert "default-src 'self'" in response.headers["content-security-policy"]


def test_readyz_returns_degraded_when_frontend_dist_is_missing(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    missing_frontend_dist = temp_dir / f"{uuid.uuid4()}-missing-dist"

    monkeypatch.setenv("AB_FRONTEND_DIST_PATH", str(missing_frontend_dist))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "true")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.get("/readyz")

    assert response.status_code == 503
    payload = response.json()
    assert payload["status"] == "degraded"
    assert any(check["name"] == "frontend_dist" and check["ok"] is False for check in payload["checks"])
    assert any(check["name"] == "sqlite_schema_version" and check["ok"] is True for check in payload["checks"])
    assert any(check["name"] == "sqlite_journal_mode" and check["ok"] is True for check in payload["checks"])
    assert any(check["name"] == "sqlite_write_probe" and check["ok"] is True for check in payload["checks"])
    get_settings.cache_clear()


def test_api_token_protects_runtime_and_api_routes(monkeypatch) -> None:
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        health_response = client.get("/health")
        unauthorized_diagnostics = client.get("/api/v1/diagnostics")
        unauthorized_readyz = client.get("/readyz")
        authorized_diagnostics = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": "Bearer super-secret-token"},
        )
        api_key_diagnostics = client.get(
            "/api/v1/diagnostics",
            headers={"X-API-Key": "super-secret-token"},
        )

    assert health_response.status_code == 200
    assert unauthorized_diagnostics.status_code == 401
    assert unauthorized_diagnostics.json()["detail"] == "Unauthorized"
    assert unauthorized_diagnostics.json()["error_code"] == "unauthorized"
    assert unauthorized_readyz.status_code == 401
    assert authorized_diagnostics.status_code == 200
    assert authorized_diagnostics.json()["auth"]["enabled"] is True
    assert authorized_diagnostics.json()["auth"]["mode"] == "token"
    assert authorized_diagnostics.json()["auth"]["write_enabled"] is True
    assert authorized_diagnostics.json()["auth"]["readonly_enabled"] is False
    assert api_key_diagnostics.status_code == 200
    get_settings.cache_clear()


def test_docs_and_openapi_remain_public_when_auth_is_enabled(monkeypatch) -> None:
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        docs_response = client.get("/docs")
        redoc_response = client.get("/redoc")
        openapi_response = client.get("/openapi.json")

    assert docs_response.status_code == 200
    assert "Swagger UI" in docs_response.text
    assert redoc_response.status_code == 200
    assert "AB Test Research Designer API" in redoc_response.text
    assert openapi_response.status_code == 200
    payload = openapi_response.json()
    assert payload["info"]["title"] == "AB Test Research Designer API"
    assert payload["info"]["version"] == "1.0.0"
    assert payload["info"]["description"]
    assert payload["info"]["contact"]
    assert payload["info"]["license"]
    get_settings.cache_clear()


def test_api_rate_limiting_rejects_excess_requests(monkeypatch) -> None:
    monkeypatch.setenv("AB_RATE_LIMIT_ENABLED", "true")
    monkeypatch.setenv("AB_RATE_LIMIT_REQUESTS", "2")
    monkeypatch.setenv("AB_RATE_LIMIT_WINDOW_SECONDS", "60")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        first = client.get("/api/v1/diagnostics")
        second = client.get("/api/v1/diagnostics")
        third = client.get("/api/v1/diagnostics")

    assert first.status_code == 200
    assert second.status_code == 200
    assert third.status_code == 429
    assert third.json()["detail"] == "Too many requests"
    assert third.json()["error_code"] == "rate_limited"
    assert int(third.headers["retry-after"]) >= 1
    assert third.headers["x-content-type-options"] == "nosniff"
    get_settings.cache_clear()


def test_readonly_api_token_allows_safe_requests_but_blocks_mutations(monkeypatch) -> None:
    monkeypatch.setenv("AB_READONLY_API_TOKEN", "readonly-secret")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        diagnostics_response = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": "Bearer readonly-secret"},
        )
        project_list_response = client.get(
            "/api/v1/projects",
            headers={"Authorization": "Bearer readonly-secret"},
        )
        forbidden_calculation = client.post(
            "/api/v1/calculate",
            headers={"Authorization": "Bearer readonly-secret"},
            json={
                "metric_type": "binary",
                "baseline_value": 0.042,
                "mde_pct": 5,
                "alpha": 0.05,
                "power": 0.8,
                "expected_daily_traffic": 12000,
                "audience_share_in_test": 0.6,
                "traffic_split": [50, 50],
                "variants_count": 2,
            },
        )

    assert diagnostics_response.status_code == 200
    assert diagnostics_response.json()["auth"]["mode"] == "readonly"
    assert diagnostics_response.json()["auth"]["write_enabled"] is False
    assert diagnostics_response.json()["auth"]["readonly_enabled"] is True
    assert project_list_response.status_code == 200
    assert forbidden_calculation.status_code == 403
    assert forbidden_calculation.json()["detail"] == "Forbidden"
    assert forbidden_calculation.json()["error_code"] == "forbidden"
    get_settings.cache_clear()


def test_auth_failure_throttling_rejects_repeated_invalid_tokens(monkeypatch) -> None:
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_RATE_LIMIT_ENABLED", "true")
    monkeypatch.setenv("AB_RATE_LIMIT_REQUESTS", "50")
    monkeypatch.setenv("AB_AUTH_FAILURE_LIMIT", "2")
    monkeypatch.setenv("AB_AUTH_FAILURE_WINDOW_SECONDS", "60")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        first = client.get("/api/v1/diagnostics", headers={"Authorization": "Bearer wrong-token"})
        second = client.get("/api/v1/diagnostics", headers={"Authorization": "Bearer wrong-token"})
        third = client.get("/api/v1/diagnostics", headers={"Authorization": "Bearer wrong-token"})
        authorized = client.get("/api/v1/diagnostics", headers={"Authorization": "Bearer super-secret-token"})

    assert first.status_code == 401
    assert second.status_code == 401
    assert first.json()["error_code"] == "unauthorized"
    assert third.status_code == 429
    assert third.json()["detail"] == "Too many unauthorized requests"
    assert third.json()["error_code"] == "auth_rate_limited"
    assert int(third.headers["retry-after"]) >= 1
    assert third.headers["x-frame-options"] == "DENY"
    assert authorized.status_code == 200
    get_settings.cache_clear()


def test_dual_token_auth_mode_reports_both_scopes(monkeypatch) -> None:
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_READONLY_API_TOKEN", "readonly-secret")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        diagnostics_response = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": "Bearer super-secret-token"},
        )

    assert diagnostics_response.status_code == 200
    assert diagnostics_response.json()["auth"]["mode"] == "dual_token"
    assert diagnostics_response.json()["auth"]["write_enabled"] is True
    assert diagnostics_response.json()["auth"]["readonly_enabled"] is True
    get_settings.cache_clear()


def test_diagnostics_runtime_counters_track_errors_and_auth_rejections(monkeypatch) -> None:
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_READONLY_API_TOKEN", "readonly-secret")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        client.get("/health")
        client.get("/api/v1/diagnostics")
        client.get("/api/v1/projects", headers={"Authorization": "Bearer readonly-secret"})
        client.post(
            "/api/v1/calculate",
            headers={"Authorization": "Bearer readonly-secret"},
            json={
                "metric_type": "binary",
                "baseline_value": 0.042,
                "mde_pct": 5,
                "alpha": 0.05,
                "power": 0.8,
                "expected_daily_traffic": 12000,
                "audience_share_in_test": 0.6,
                "traffic_split": [50, 50],
                "variants_count": 2,
            },
        )
        diagnostics_response = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": "Bearer super-secret-token"},
        )

    assert diagnostics_response.status_code == 200
    runtime = diagnostics_response.json()["runtime"]
    assert runtime["total_requests"] >= 4
    assert runtime["success_responses"] >= 2
    assert runtime["client_error_responses"] >= 1
    assert runtime["server_error_responses"] == 0
    assert runtime["auth_rejections"] >= 1
    assert runtime["rate_limited_responses"] == 0
    assert runtime["request_body_rejections"] == 0
    assert runtime["last_error_code"] == "forbidden"
    assert runtime["last_error_at"] is not None
    assert runtime["last_request_at"] is not None
    get_settings.cache_clear()


def test_database_api_keys_require_auth_and_enforce_scope(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"
    repository = ProjectRepository(str(db_path))
    write_key = repository.create_api_key(name="Write key", scope="write")["plaintext_key"]
    read_key = repository.create_api_key(name="Read key", scope="read")["plaintext_key"]

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        unauthorized = client.get("/api/v1/diagnostics")
        write_response = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": f"Bearer {write_key}"},
        )
        readonly_projects = client.get(
            "/api/v1/projects",
            headers={"Authorization": f"Bearer {read_key}"},
        )
        forbidden_calculation = client.post(
            "/api/v1/calculate",
            headers={"Authorization": f"Bearer {read_key}"},
            json={
                "metric_type": "binary",
                "baseline_value": 0.042,
                "mde_pct": 5,
                "alpha": 0.05,
                "power": 0.8,
                "expected_daily_traffic": 12000,
                "audience_share_in_test": 0.6,
                "traffic_split": [50, 50],
                "variants_count": 2,
            },
        )

    assert unauthorized.status_code == 401
    assert write_response.status_code == 200
    assert write_response.json()["auth"]["session_scope"] == "write"
    assert write_response.json()["auth"]["session_can_write"] is True
    assert readonly_projects.status_code == 200
    assert forbidden_calculation.status_code == 403
    assert forbidden_calculation.json()["error_code"] == "forbidden"
    get_settings.cache_clear()


def test_revoked_database_api_key_is_rejected(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"
    repository = ProjectRepository(str(db_path))
    created = repository.create_api_key(name="Temporary key", scope="write")
    repository.revoke_api_key(created["id"])

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": f"Bearer {created['plaintext_key']}"},
        )

    assert response.status_code == 401
    assert response.json()["error_code"] == "unauthorized"
    get_settings.cache_clear()


def test_per_key_rate_limits_use_independent_buckets(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"
    repository = ProjectRepository(str(db_path))
    first_key = repository.create_api_key(name="Partner A", scope="write")
    second_key = repository.create_api_key(name="Partner B", scope="write")

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_RATE_LIMIT_ENABLED", "true")
    monkeypatch.setenv("AB_RATE_LIMIT_REQUESTS", "1")
    monkeypatch.setenv("AB_RATE_LIMIT_WINDOW_SECONDS", "60")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        first_partner_first = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": f"Bearer {first_key['plaintext_key']}"},
        )
        second_partner_first = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": f"Bearer {second_key['plaintext_key']}"},
        )
        first_partner_second = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": f"Bearer {first_key['plaintext_key']}"},
        )

    assert first_partner_first.status_code == 200
    assert second_partner_first.status_code == 200
    assert first_partner_second.status_code == 429
    assert first_partner_second.json()["error_code"] == "rate_limited"
    get_settings.cache_clear()


def test_api_key_usage_events_are_filterable_by_key_id(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"
    repository = ProjectRepository(str(db_path))
    created = repository.create_api_key(name="Audited key", scope="write")

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        diagnostics = client.get(
            "/api/v1/diagnostics",
            headers={"Authorization": f"Bearer {created['plaintext_key']}"},
        )
        audit = client.get(
            "/api/v1/audit",
            params={"action": "api_key_used", "key_id": created["id"]},
            headers={"Authorization": f"Bearer {created['plaintext_key']}"},
        )

    assert diagnostics.status_code == 200
    assert audit.status_code == 200
    assert audit.json()["entries"]
    assert all(entry["key_id"] == created["id"] for entry in audit.json()["entries"])
    assert any(entry["action"] == "api_key_used" for entry in audit.json()["entries"])
    get_settings.cache_clear()


def test_webhook_routes_require_admin_auth(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"
    repository = ProjectRepository(str(db_path))
    write_key = repository.create_api_key(name="Write key", scope="write")["plaintext_key"]

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_ADMIN_TOKEN", "admin-secret-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    payload = {
        "name": "Partner alerts",
        "target_url": "https://example.com/webhook",
        "secret": "top-secret",
        "format": "generic",
        "event_filter": ["api_key_created"],
        "scope": "global",
    }

    with TestClient(create_app()) as client:
        unauthorized = client.get("/api/v1/webhooks")
        forbidden = client.post(
            "/api/v1/webhooks",
            headers={"Authorization": f"Bearer {write_key}"},
            json=payload,
        )
        authorized = client.post(
            "/api/v1/webhooks",
            headers={"Authorization": "Bearer admin-secret-token"},
            json=payload,
        )

    assert unauthorized.status_code == 401
    assert forbidden.status_code == 403
    assert authorized.status_code == 200
    assert authorized.json()["secret"] == "top-secret"
    get_settings.cache_clear()


def test_webhook_routes_support_crud_and_hide_secret_after_create(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_ADMIN_TOKEN", "admin-secret-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        created = client.post(
            "/api/v1/webhooks",
            headers={"Authorization": "Bearer admin-secret-token"},
            json={
                "name": "Partner alerts",
                "target_url": "https://example.com/webhook",
                "secret": "top-secret",
                "format": "generic",
                "event_filter": ["api_key_created"],
                "scope": "global",
            },
        )
        assert created.status_code == 200
        webhook_id = created.json()["id"]

        listed = client.get(
            "/api/v1/webhooks",
            headers={"Authorization": "Bearer admin-secret-token"},
        )
        fetched = client.get(
            f"/api/v1/webhooks/{webhook_id}",
            headers={"Authorization": "Bearer admin-secret-token"},
        )
        updated = client.patch(
            f"/api/v1/webhooks/{webhook_id}",
            headers={"Authorization": "Bearer admin-secret-token"},
            json={
                "enabled": False,
                "event_filter": ["api_key_revoked"],
                "target_url": "https://example.com/updated",
            },
        )
        deliveries = client.get(
            f"/api/v1/webhooks/{webhook_id}/deliveries",
            headers={"Authorization": "Bearer admin-secret-token"},
        )
        deleted = client.delete(
            f"/api/v1/webhooks/{webhook_id}",
            headers={"Authorization": "Bearer admin-secret-token"},
        )

    assert listed.status_code == 200
    assert fetched.status_code == 200
    assert updated.status_code == 200
    assert deliveries.status_code == 200
    assert deleted.status_code == 200
    assert listed.json()["subscriptions"][0]["secret"] is None
    assert fetched.json()["secret"] is None
    assert updated.json()["enabled"] is False
    assert updated.json()["event_filter"] == ["api_key_revoked"]
    assert updated.json()["target_url"] == "https://example.com/updated"
    assert deliveries.json()["deliveries"] == []
    assert deleted.json() == {"id": webhook_id, "deleted": True}
    get_settings.cache_clear()

def test_api_token_does_not_break_cors_preflight(monkeypatch) -> None:
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.options(
            "/api/v1/calculate",
            headers={
                "Origin": "http://127.0.0.1:5173",
                "Access-Control-Request-Method": "POST",
                "Access-Control-Request-Headers": "authorization,content-type",
            },
        )

    assert response.status_code == 200
    assert response.headers["access-control-allow-origin"] == "http://127.0.0.1:5173"
    get_settings.cache_clear()


def test_project_report_pdf_endpoint_returns_pdf_document(monkeypatch) -> None:
    from io import BytesIO

    from pypdf import PdfReader

    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        created = client.post("/api/v1/projects", json=_full_payload())
        assert created.status_code == 200
        project_id = created.json()["id"]

        analysis_saved = client.post(
            f"/api/v1/projects/{project_id}/analysis",
            json=_saved_analysis_payload(
                total_sample_size=320,
                estimated_duration_days=14,
                executive_summary="Checkout summary",
            ),
        )
        assert analysis_saved.status_code == 200

        response = client.get(f"/api/v1/projects/{project_id}/report/pdf")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/pdf")
    assert "checkout-redesign-report.pdf" in response.headers["content-disposition"]
    reader = PdfReader(BytesIO(response.content))
    assert len(reader.pages) == 4
    extracted = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Experiment Summary" in extracted
    assert "Sample Size Results" in extracted
    assert "Power Curve" in extracted
    assert "Warnings & Recommendations" in extracted
    get_settings.cache_clear()


def test_project_report_pdf_endpoint_returns_404_for_missing_project(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.get("/api/v1/projects/missing-project/report/pdf")

    assert response.status_code == 404
    assert response.json()["detail"] == "Project not found"
    get_settings.cache_clear()


def test_project_report_pdf_endpoint_requires_auth(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.get("/api/v1/projects/missing-project/report/pdf")

    assert response.status_code == 401
    assert response.json()["detail"] == "Unauthorized"
    get_settings.cache_clear()


def test_projects_list_endpoint_supports_search_and_metadata(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    hypothesis_match_payload = _full_payload()
    hypothesis_match_payload["project"]["project_name"] = "Pricing experiment"
    hypothesis_match_payload["hypothesis"]["hypothesis_statement"] = "Checkout velocity will improve."
    hypothesis_match_payload["metrics"]["metric_type"] = "continuous"
    hypothesis_match_payload["metrics"]["primary_metric_name"] = "avg_order_value"
    hypothesis_match_payload["metrics"]["baseline_value"] = 45.0
    hypothesis_match_payload["metrics"]["std_dev"] = 12.0

    with TestClient(create_app()) as client:
        created_checkout = client.post("/api/v1/projects", json=_full_payload())
        created_pricing = client.post("/api/v1/projects", json=hypothesis_match_payload)
        assert created_checkout.status_code == 200
        assert created_pricing.status_code == 200

        response = client.get("/api/v1/projects", params={"q": "checkout", "limit": 1, "offset": 0})

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert payload["limit"] == 1
    assert payload["offset"] == 0
    assert payload["has_more"] is True
    assert len(payload["projects"]) == 1
    assert payload["projects"][0]["project_name"] in {"Checkout redesign", "Pricing experiment"}
    get_settings.cache_clear()


def test_projects_list_endpoint_filters_archived_projects(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    archived_payload = _full_payload()
    archived_payload["project"]["project_name"] = "Archived checkout"

    with TestClient(create_app()) as client:
        active_project = client.post("/api/v1/projects", json=_full_payload())
        archived_project = client.post("/api/v1/projects", json=archived_payload)
        assert active_project.status_code == 200
        assert archived_project.status_code == 200
        client.post(f"/api/v1/projects/{archived_project.json()['id']}/archive")

        response = client.get("/api/v1/projects", params={"status": "archived"})

    assert response.status_code == 200
    assert response.json()["total"] == 1
    assert response.json()["projects"][0]["project_name"] == "Archived checkout"
    assert response.json()["projects"][0]["is_archived"] is True
    get_settings.cache_clear()


def test_projects_list_endpoint_filters_and_sorts_by_name(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    alpha_payload = _full_payload()
    alpha_payload["project"]["project_name"] = "Alpha checkout"

    zeta_payload = _full_payload()
    zeta_payload["project"]["project_name"] = "Zeta checkout"

    continuous_payload = _continuous_full_payload()
    continuous_payload["project"]["project_name"] = "Mid pricing"

    with TestClient(create_app()) as client:
        client.post("/api/v1/projects", json=zeta_payload)
        client.post("/api/v1/projects", json=continuous_payload)
        client.post("/api/v1/projects", json=alpha_payload)

        response = client.get(
            "/api/v1/projects",
            params={
                "status": "all",
                "metric_type": "binary",
                "sort_by": "name",
                "sort_dir": "asc",
            },
        )

    assert response.status_code == 200
    names = [project["project_name"] for project in response.json()["projects"]]
    assert names == ["Alpha checkout", "Zeta checkout"]
    get_settings.cache_clear()


def test_projects_list_endpoint_sorts_by_duration_days(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    fast_payload = _full_payload()
    fast_payload["project"]["project_name"] = "Fast checkout"
    slow_payload = _full_payload()
    slow_payload["project"]["project_name"] = "Slow checkout"

    with TestClient(create_app()) as client:
        fast = client.post("/api/v1/projects", json=fast_payload)
        slow = client.post("/api/v1/projects", json=slow_payload)
        assert fast.status_code == 200
        assert slow.status_code == 200

        client.post(
            f"/api/v1/projects/{fast.json()['id']}/analysis",
            json=_saved_analysis_payload(total_sample_size=200, estimated_duration_days=6, executive_summary="Fast"),
        )
        client.post(
            f"/api/v1/projects/{slow.json()['id']}/analysis",
            json=_saved_analysis_payload(total_sample_size=500, estimated_duration_days=18, executive_summary="Slow"),
        )

        response = client.get(
            "/api/v1/projects",
            params={"status": "all", "sort_by": "duration_days", "sort_dir": "asc"},
        )

    assert response.status_code == 200
    projects = response.json()["projects"]
    assert [project["project_name"] for project in projects[:2]] == ["Fast checkout", "Slow checkout"]
    assert projects[0]["duration_days"] == 6
    assert projects[1]["duration_days"] == 18
    get_settings.cache_clear()


def test_audit_log_tracks_create_update_archive_delete_and_sorts_newest_first(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    updated_payload = _full_payload()
    updated_payload["project"]["project_name"] = "Checkout redesign v2"

    with TestClient(create_app()) as client:
        created = client.post("/api/v1/projects", json=_full_payload())
        assert created.status_code == 200
        project_id = created.json()["id"]

        updated = client.put(f"/api/v1/projects/{project_id}", json=updated_payload)
        assert updated.status_code == 200

        archived = client.post(f"/api/v1/projects/{project_id}/archive")
        assert archived.status_code == 200

        deleted = client.delete(f"/api/v1/projects/{project_id}")
        assert deleted.status_code == 200

        response = client.get("/api/v1/audit")

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 4
    assert [entry["action"] for entry in payload["entries"]] == [
        "project.delete",
        "project.archive",
        "project.update",
        "project.create",
    ]
    get_settings.cache_clear()


def test_audit_log_records_payload_diff_for_project_updates(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    updated_payload = _full_payload()
    updated_payload["hypothesis"]["hypothesis_statement"] = "A shorter checkout will reduce friction."

    with TestClient(create_app()) as client:
        created = client.post("/api/v1/projects", json=_full_payload())
        assert created.status_code == 200
        project_id = created.json()["id"]

        updated = client.put(f"/api/v1/projects/{project_id}", json=updated_payload)
        assert updated.status_code == 200

        response = client.get("/api/v1/audit", params={"action": "project.update"})

    assert response.status_code == 200
    entries = response.json()["entries"]
    assert len(entries) == 1
    assert entries[0]["action"] == "project.update"
    assert entries[0]["payload_diff"]["hypothesis.hypothesis_statement"] == [
        "If we simplify checkout, conversion will increase.",
        "A shorter checkout will reduce friction.",
    ]
    get_settings.cache_clear()


def test_audit_log_filters_by_project_id(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    second_payload = _full_payload()
    second_payload["project"]["project_name"] = "Pricing test"

    with TestClient(create_app()) as client:
        first = client.post("/api/v1/projects", json=_full_payload())
        second = client.post("/api/v1/projects", json=second_payload)
        assert first.status_code == 200
        assert second.status_code == 200

        response = client.get("/api/v1/audit", params={"project_id": second.json()["id"]})

    assert response.status_code == 200
    entries = response.json()["entries"]
    assert len(entries) == 1
    assert entries[0]["project_id"] == second.json()["id"]
    assert entries[0]["project_name"] == "Pricing test"
    get_settings.cache_clear()


def test_audit_log_export_returns_csv(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        created = client.post("/api/v1/projects", json=_full_payload())
        assert created.status_code == 200

        response = client.get("/api/v1/audit/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "action,project_id,project_name,actor" in response.text
    assert "project.create" in response.text
    get_settings.cache_clear()


def test_audit_log_requires_auth_and_export_is_forbidden_for_readonly_tokens(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_READONLY_API_TOKEN", "readonly-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        unauthorized = client.get("/api/v1/audit")
        readonly_list = client.get(
            "/api/v1/audit",
            headers={"Authorization": "Bearer readonly-token"},
        )
        readonly_export = client.get(
            "/api/v1/audit/export",
            headers={"Authorization": "Bearer readonly-token"},
        )

    assert unauthorized.status_code == 401
    assert readonly_list.status_code == 200
    assert readonly_export.status_code == 403
    get_settings.cache_clear()


def test_project_report_csv_endpoint_returns_multisection_csv(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        created = client.post("/api/v1/projects", json=_full_payload())
        assert created.status_code == 200
        project_id = created.json()["id"]

        analysis_saved = client.post(
            f"/api/v1/projects/{project_id}/analysis",
            json=_saved_analysis_payload(total_sample_size=320, estimated_duration_days=14),
        )
        assert analysis_saved.status_code == 200

        response = client.get(f"/api/v1/projects/{project_id}/report/csv")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "# Sample Size Results" in response.text
    assert "# Sensitivity Analysis" in response.text
    assert "# Guardrail Metrics" in response.text
    get_settings.cache_clear()


def test_project_report_xlsx_endpoint_returns_workbook_with_four_sheets(monkeypatch) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        created = client.post("/api/v1/projects", json=_full_payload())
        assert created.status_code == 200
        project_id = created.json()["id"]

        analysis_saved = client.post(
            f"/api/v1/projects/{project_id}/analysis",
            json=_saved_analysis_payload(total_sample_size=320, estimated_duration_days=14),
        )
        assert analysis_saved.status_code == 200

        response = client.get(f"/api/v1/projects/{project_id}/report/xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Sensitivity", "Guardrails", "Raw Inputs"]
    get_settings.cache_clear()


def test_project_report_data_export_returns_404_for_missing_project(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        csv_response = client.get("/api/v1/projects/missing-project/report/csv")
        xlsx_response = client.get("/api/v1/projects/missing-project/report/xlsx")

    assert csv_response.status_code == 404
    assert xlsx_response.status_code == 404
    get_settings.cache_clear()


def test_project_report_data_export_requires_auth(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_API_TOKEN", "super-secret-token")
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        csv_response = client.get("/api/v1/projects/missing-project/report/csv")
        xlsx_response = client.get("/api/v1/projects/missing-project/report/xlsx")

    assert csv_response.status_code == 401
    assert xlsx_response.status_code == 401
    get_settings.cache_clear()


def test_templates_list_returns_five_builtins_on_fresh_install(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        response = client.get("/api/v1/templates")

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 5
    assert len(payload["templates"]) == 5
    assert all(template["built_in"] is True for template in payload["templates"])
    get_settings.cache_clear()


def test_template_use_returns_payload_and_increments_usage_count(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        listed = client.get("/api/v1/templates")
        assert listed.status_code == 200
        template_id = listed.json()["templates"][0]["id"]

        used = client.post(f"/api/v1/templates/{template_id}/use")
        assert used.status_code == 200

        refreshed = client.get(f"/api/v1/templates/{template_id}")

    assert refreshed.status_code == 200
    used_payload = used.json()
    assert used_payload["payload"]["project"]["project_name"] == ""
    assert used_payload["usage_count"] == 1
    assert refreshed.json()["usage_count"] == 1
    get_settings.cache_clear()


def test_templates_create_saves_user_defined_template(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    payload = _full_payload()
    payload["project"]["project_name"] = "Saved template draft"

    with TestClient(create_app()) as client:
        response = client.post(
            "/api/v1/templates",
            json={
                "name": "Saved checkout flow",
                "description": "Team template for checkout experiments.",
                "category": "Revenue",
                "tags": ["binary", "checkout"],
                "payload": payload,
            },
        )

    assert response.status_code == 200
    created = response.json()
    assert created["built_in"] is False
    assert created["name"] == "Saved checkout flow"
    assert created["payload"]["project"]["project_name"] == "Saved template draft"
    get_settings.cache_clear()


def test_templates_delete_rejects_builtin_templates(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        listed = client.get("/api/v1/templates")
        assert listed.status_code == 200
        template_id = listed.json()["templates"][0]["id"]
        response = client.delete(f"/api/v1/templates/{template_id}")

    assert response.status_code == 403
    get_settings.cache_clear()


def test_templates_delete_removes_user_template(monkeypatch) -> None:
    temp_dir = Path(__file__).resolve().parent / ".tmp"
    temp_dir.mkdir(exist_ok=True)
    db_path = temp_dir / f"{uuid.uuid4()}.sqlite3"

    monkeypatch.setenv("AB_DB_PATH", str(db_path))
    monkeypatch.setenv("AB_SERVE_FRONTEND_DIST", "false")
    get_settings.cache_clear()

    with TestClient(create_app()) as client:
        created = client.post(
            "/api/v1/templates",
            json={
                "name": "My pricing template",
                "description": "Reusable pricing experiment baseline.",
                "category": "Revenue",
                "tags": ["pricing"],
                "payload": _full_payload(),
            },
        )
        assert created.status_code == 200
        template_id = created.json()["id"]

        deleted = client.delete(f"/api/v1/templates/{template_id}")
        listed = client.get("/api/v1/templates")

    assert deleted.status_code == 200
    assert deleted.json() == {"id": template_id, "deleted": True}
    assert all(template["id"] != template_id for template in listed.json()["templates"])
    get_settings.cache_clear()
