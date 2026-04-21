import csv
from datetime import datetime, timezone
from io import BytesIO, StringIO
from html import escape
import json
import math
from statistics import NormalDist
from typing import Any

from openpyxl import Workbook

from app.backend.app.i18n import translate
from app.backend.app.schemas.api import StandaloneExportRequest
from app.backend.app.services.pdf_service import _build_sensitivity_rows


def _list_to_markdown(items: list[str]) -> str:
    if not items:
        return "- None"
    return "\n".join(f"- {item}" for item in items)


def _as_html(value: object) -> str:
    return escape(str(value))


def _format_metric_value(value: object) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}"
        return f"{value:.2f}"
    return _as_html(value)


def _render_power_curve_svg(calculation: dict[str, Any]) -> str:
    summary = calculation.get("calculation_summary", {})
    results = calculation.get("results", {})
    sample_size = results.get("sample_size_per_variant")
    alpha = summary.get("alpha")
    current_power = summary.get("power")

    if not isinstance(sample_size, int) or sample_size <= 0:
        return ""
    if not isinstance(alpha, (int, float)) or not 0 < alpha < 1:
        return ""
    if not isinstance(current_power, (int, float)) or not 0 < current_power < 1:
        return ""

    powers = sorted({0.7, 0.8, 0.9, 0.95, round(float(current_power), 4)})
    dist = NormalDist()
    z_alpha = dist.inv_cdf(1 - (float(alpha) / 2))
    z_current = dist.inv_cdf(float(current_power))
    estimated = [
        (
            power,
            max(
                1,
                round(
                    sample_size
                    * (((z_alpha + dist.inv_cdf(power)) / (z_alpha + z_current)) ** 2)
                ),
            ),
        )
        for power in powers
    ]
    values = [item[1] for item in estimated]
    min_value = min(values)
    max_value = max(values)
    width = 640
    height = 280
    padding_left = 64
    padding_right = 20
    padding_top = 24
    padding_bottom = 48
    chart_width = width - padding_left - padding_right
    chart_height = height - padding_top - padding_bottom

    def scale_x(power: float) -> float:
        span = powers[-1] - powers[0] or 1
        return padding_left + ((power - powers[0]) / span) * chart_width

    def scale_y(value: int) -> float:
        span = max_value - min_value or 1
        return padding_top + chart_height - ((value - min_value) / span) * chart_height

    points = " ".join(f"{scale_x(power):.1f},{scale_y(value):.1f}" for power, value in estimated)
    x_ticks = "".join(
        f'<g><line x1="{scale_x(power):.1f}" y1="{padding_top}" x2="{scale_x(power):.1f}" y2="{padding_top + chart_height}" '
        f'stroke="#e2e8f0" stroke-dasharray="4 4" />'
        f'<text x="{scale_x(power):.1f}" y="{height - 14}" text-anchor="middle" fill="#64748b" font-size="12">{power:.2f}</text></g>'
        for power in powers
    )
    y_ticks = "".join(
        f'<g><line x1="{padding_left}" y1="{scale_y(value):.1f}" x2="{width - padding_right}" y2="{scale_y(value):.1f}" '
        f'stroke="#e2e8f0" stroke-dasharray="4 4" />'
        f'<text x="{padding_left - 12}" y="{scale_y(value) + 4:.1f}" text-anchor="end" fill="#64748b" font-size="12">{value:,}</text></g>'
        for value in sorted(set(values))
    )
    dots = "".join(
        f'<g><circle cx="{scale_x(power):.1f}" cy="{scale_y(value):.1f}" r="5" fill="#0d9488" />'
        f'<text x="{scale_x(power):.1f}" y="{scale_y(value) - 12:.1f}" text-anchor="middle" fill="#0f172a" font-size="11">{value:,}</text></g>'
        for power, value in estimated
    )

    return f"""<div class="section">
  <h2>Power Curve Approximation</h2>
  <div class="svg-shell">
    <svg viewBox="0 0 {width} {height}" role="img" aria-label="Power curve approximation">
      <rect x="0" y="0" width="{width}" height="{height}" rx="16" fill="#ffffff" />
      {y_ticks}
      {x_ticks}
      <line x1="{padding_left}" y1="{padding_top + chart_height}" x2="{width - padding_right}" y2="{padding_top + chart_height}" stroke="#94a3b8" />
      <line x1="{padding_left}" y1="{padding_top}" x2="{padding_left}" y2="{padding_top + chart_height}" stroke="#94a3b8" />
      <polyline fill="none" stroke="#0d9488" stroke-width="3" points="{points}" />
      {dots}
      <text x="{width / 2:.1f}" y="{height - 4}" text-anchor="middle" fill="#475569" font-size="12">Power target</text>
      <text x="18" y="{height / 2:.1f}" text-anchor="middle" fill="#475569" font-size="12" transform="rotate(-90 18 {height / 2:.1f})">Sample size / variant</text>
    </svg>
  </div>
</div>"""


def _duration_color(duration: float) -> str:
    if duration <= 7:
        return "#ccfbf1"
    if duration <= 14:
        return "#fef3c7"
    return "#fee2e2"


def _render_sensitivity_svg(data: dict[str, Any] | None) -> str:
    if not isinstance(data, dict):
        return ""

    cells = data.get("cells", [])
    if not isinstance(cells, list) or not cells:
        return ""

    current_mde = data.get("current_mde")
    current_power = data.get("current_power")
    mde_values = sorted(
        {
            float(cell["mde"])
            for cell in cells
            if isinstance(cell, dict) and isinstance(cell.get("mde"), (int, float))
        }
    )
    power_values = sorted(
        {
            float(cell["power"])
            for cell in cells
            if isinstance(cell, dict) and isinstance(cell.get("power"), (int, float))
        }
    )
    if not mde_values or not power_values:
        return ""

    cell_map = {
        (float(cell["mde"]), float(cell["power"])): cell
        for cell in cells
        if isinstance(cell, dict)
        and isinstance(cell.get("mde"), (int, float))
        and isinstance(cell.get("power"), (int, float))
    }
    cell_width = 116
    cell_height = 54
    header_width = 90
    header_height = 48
    width = header_width + (len(power_values) * cell_width) + 24
    height = header_height + (len(mde_values) * cell_height) + 62
    parts = []

    for index, power in enumerate(power_values):
        x = header_width + (index * cell_width) + (cell_width / 2)
        parts.append(
            f'<text x="{x:.1f}" y="30" text-anchor="middle" fill="#475569" font-size="12" font-weight="600">{power:.2f}</text>'
        )

    for index, mde in enumerate(mde_values):
        y = header_height + (index * cell_height)
        parts.append(
            f'<text x="{header_width - 12}" y="{y + (cell_height / 2) + 5:.1f}" text-anchor="end" fill="#475569" font-size="12" font-weight="600">{mde:g}</text>'
        )
        for power_index, power in enumerate(power_values):
            cell = cell_map.get((mde, power))
            x = header_width + (power_index * cell_width)
            if cell is None:
                parts.append(
                    f'<rect x="{x}" y="{y}" width="{cell_width}" height="{cell_height}" rx="10" fill="#f8fafc" stroke="#e2e8f0" />'
                )
                parts.append(
                    f'<text x="{x + (cell_width / 2):.1f}" y="{y + (cell_height / 2) + 4:.1f}" text-anchor="middle" fill="#94a3b8" font-size="12">n/a</text>'
                )
                continue

            duration = float(cell.get("duration_days", 0))
            stroke = "#0d9488" if mde == current_mde and power == current_power else "#e2e8f0"
            stroke_width = "3" if mde == current_mde and power == current_power else "1.5"
            sample = cell.get("sample_size_per_variant")
            parts.append(
                f'<rect x="{x}" y="{y}" width="{cell_width}" height="{cell_height}" rx="10" fill="{_duration_color(duration)}" stroke="{stroke}" stroke-width="{stroke_width}" />'
            )
            parts.append(
                f'<text x="{x + (cell_width / 2):.1f}" y="{y + 23:.1f}" text-anchor="middle" fill="#0f172a" font-size="14" font-weight="700">{math.ceil(duration)}d</text>'
            )
            parts.append(
                f'<text x="{x + (cell_width / 2):.1f}" y="{y + 40:.1f}" text-anchor="middle" fill="#475569" font-size="11">n={_format_metric_value(sample)}</text>'
            )

    parts.append(
        '<text x="20" y="30" fill="#475569" font-size="12" font-weight="600">MDE</text>'
    )
    parts.append(
        f'<text x="{width / 2:.1f}" y="{height - 16}" text-anchor="middle" fill="#475569" font-size="12">Power target</text>'
    )
    parts.append(
        f'<text x="20" y="{height / 2:.1f}" text-anchor="middle" fill="#475569" font-size="12" transform="rotate(-90 20 {height / 2:.1f})">Duration</text>'
    )

    return f"""<div class="section">
  <h2>Sensitivity Table</h2>
  <div class="svg-shell">
    <svg viewBox="0 0 {width} {height}" role="img" aria-label="Sensitivity table">
      <rect x="0" y="0" width="{width}" height="{height}" rx="16" fill="#ffffff" />
      {''.join(parts)}
    </svg>
  </div>
</div>"""


def _render_hypothesis_section(hypothesis: str | None) -> str:
    if not hypothesis:
        return ""
    return f"""<div class="section">
  <h2>{translate("export.standalone.hypothesis")}</h2>
  <p>{_as_html(hypothesis)}</p>
</div>"""


def _render_design_section(design: dict[str, Any]) -> str:
    variants = design.get("variants", [])
    variant_rows = "".join(
        f"<tr><td>{_as_html(variant.get('name', 'Variant'))}</td><td>{_as_html(variant.get('description', ''))}</td></tr>"
        for variant in variants
        if isinstance(variant, dict)
    )
    stopping_conditions = design.get("stopping_conditions", [])
    stopping_list = "".join(f"<li>{_as_html(item)}</li>" for item in stopping_conditions)
    return f"""<div class="section">
  <h2>Design Details</h2>
  <table>
    <tbody>
      <tr><th>Randomization unit</th><td>{_as_html(design.get('randomization_unit', '—'))}</td></tr>
      <tr><th>Traffic split</th><td>{_as_html(design.get('traffic_split', '—'))}</td></tr>
      <tr><th>Target audience</th><td>{_as_html(design.get('target_audience', '—'))}</td></tr>
      <tr><th>Inclusion criteria</th><td>{_as_html(design.get('inclusion_criteria', '—'))}</td></tr>
      <tr><th>Exclusion criteria</th><td>{_as_html(design.get('exclusion_criteria', '—'))}</td></tr>
      <tr><th>Recommended duration</th><td>{_as_html(design.get('recommended_duration_days', '—'))}</td></tr>
    </tbody>
  </table>
  {f'<div class="subsection"><h3>Variants</h3><table><thead><tr><th>Variant</th><th>Description</th></tr></thead><tbody>{variant_rows}</tbody></table></div>' if variant_rows else ''}
  {f'<div class="subsection"><h3>Stopping Conditions</h3><ul>{stopping_list}</ul></div>' if stopping_list else ''}
</div>"""


def _render_warnings_section(warnings: list[dict[str, Any]]) -> str:
    if not warnings:
        return (
            """<div class="section">
  <h2>{title}</h2>
  <div class="warning-low">{message}</div>
</div>"""
            .format(
                title=translate("export.standalone.warnings"),
                message=translate("export.standalone.no_warning_rules"),
            )
        )
    content = "".join(
        f'<div class="warning-{_as_html(str(warning.get("severity", "low")))}"><strong>{_as_html(warning.get("code", "warning"))}</strong><div>{_as_html(warning.get("message", ""))}</div></div>'
        for warning in warnings
    )
    return f"""<div class="section">
  <h2>{translate("export.standalone.warnings")}</h2>
  {content}
</div>"""


def _render_ai_section(ai_advice: dict[str, Any]) -> str:
    blocks = []
    for key, value in ai_advice.items():
        title = _as_html(str(key).replace("_", " ").title())
        if isinstance(value, list):
            items = "".join(f"<li>{_as_html(item)}</li>" for item in value)
            if items:
                blocks.append(f"<div class=\"subsection\"><h3>{title}</h3><ul>{items}</ul></div>")
            continue
        if value not in (None, ""):
            blocks.append(f"<div class=\"subsection\"><h3>{title}</h3><p>{_as_html(value)}</p></div>")
    if not blocks:
        return ""
    return f"""<div class="section">
  <h2>{translate("export.standalone.ai_recommendations")}</h2>
  {''.join(blocks)}
</div>"""


def _render_results_section(results: dict[str, Any]) -> str:
    cards = [
        (translate("export.standalone.observed_effect"), results.get("observed_effect")),
        (translate("export.standalone.relative_change"), f"{_format_metric_value(results.get('observed_effect_relative'))}%"),
        (translate("export.standalone.p_value"), results.get("p_value")),
        (translate("export.standalone.power_achieved"), results.get("power_achieved")),
    ]
    metric_cards = "".join(
        f'<div class="metric-card"><div class="metric-value">{_format_metric_value(value)}</div><div class="metric-label">{_as_html(label)}</div></div>'
        for label, value in cards
    )
    summary_rows = [
        (translate("export.standalone.metric_type"), results.get("metric_type")),
        (translate("export.standalone.control_rate"), results.get("control_rate")),
        (translate("export.standalone.treatment_rate"), results.get("treatment_rate")),
        (translate("export.standalone.confidence_interval"), f"{_format_metric_value(results.get('ci_lower'))} to {_format_metric_value(results.get('ci_upper'))}"),
        (translate("export.standalone.verdict"), results.get("verdict")),
        (translate("export.standalone.interpretation"), results.get("interpretation")),
    ]
    table_rows = "".join(
        f"<tr><th>{_as_html(label)}</th><td>{_as_html(value)}</td></tr>"
        for label, value in summary_rows
        if value not in (None, "")
    )
    return f"""<div class="section">
  <h2>{translate("export.standalone.post_experiment_results")}</h2>
  <div class="metric-grid">{metric_cards}</div>
  <table><tbody>{table_rows}</tbody></table>
</div>"""


def build_standalone_html(request: StandaloneExportRequest) -> str:
    calculation = request.calculation
    results = calculation.get("results", {})
    summary = calculation.get("calculation_summary", {})
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    assumptions = calculation.get("assumptions", [])
    assumptions_html = "".join(f"<li>{_as_html(item)}</li>" for item in assumptions)
    key_metrics = [
        (translate("export.standalone.users_per_variant"), results.get("sample_size_per_variant")),
        (translate("export.standalone.days_estimated"), results.get("estimated_duration_days")),
        (translate("export.standalone.total_sample"), results.get("total_sample_size")),
        (translate("export.standalone.alpha"), summary.get("alpha")),
    ]
    metric_cards = "".join(
        f'<div class="metric-card"><div class="metric-value">{_format_metric_value(value)}</div><div class="metric-label">{_as_html(label)}</div></div>'
        for label, value in key_metrics
    )
    power_curve_svg = _render_power_curve_svg(calculation)
    sensitivity_svg = _render_sensitivity_svg(request.sensitivity)
    warnings_section = _render_warnings_section(calculation.get("warnings", []))
    ai_section = _render_ai_section(request.ai_advice) if request.ai_advice else ""
    results_section = _render_results_section(request.results) if request.results else ""

    return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{_as_html(translate("export.standalone.title", {"project_name": request.project_name}))}</title>
    <style>
      :root {{
        --color-primary: #0d9488;
        --color-primary-soft: #ccfbf1;
        --color-border: #dbe4ea;
        --color-text: #0f172a;
        --color-muted: #64748b;
        --color-bg: #f8fafc;
        --radius: 16px;
      }}
      * {{ box-sizing: border-box; }}
      body {{ margin: 0; padding: 32px; font-family: "Segoe UI", system-ui, sans-serif; color: var(--color-text); background: linear-gradient(180deg, #f8fafc 0%, #ffffff 180px); }}
      .report {{ max-width: 980px; margin: 0 auto; }}
      .report-header {{ padding: 28px; border: 1px solid var(--color-border); border-radius: 24px; background: #ffffff; margin-bottom: 24px; }}
      .eyebrow {{ color: var(--color-primary); font-size: 0.85rem; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase; }}
      h1 {{ margin: 10px 0 8px; font-size: 2rem; line-height: 1.1; }}
      h2 {{ margin: 0 0 16px; font-size: 1.15rem; color: var(--color-primary); }}
      h3 {{ margin: 0 0 10px; font-size: 0.96rem; color: var(--color-text); }}
      p, li, td, th {{ line-height: 1.55; }}
      .meta {{ color: var(--color-muted); }}
      .section {{ margin-bottom: 20px; padding: 22px; border: 1px solid var(--color-border); border-radius: var(--radius); background: #ffffff; page-break-inside: avoid; }}
      .subsection + .subsection {{ margin-top: 18px; }}
      .metric-grid {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px; }}
      .metric-card {{ padding: 16px; border-radius: 14px; background: var(--color-bg); text-align: center; }}
      .metric-value {{ font-size: 1.65rem; font-weight: 700; color: var(--color-primary); font-variant-numeric: tabular-nums; }}
      .metric-label {{ margin-top: 6px; color: var(--color-muted); font-size: 0.86rem; }}
      .svg-shell {{ overflow-x: auto; }}
      svg {{ width: 100%; height: auto; display: block; }}
      table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
      th {{ width: 34%; text-align: left; color: var(--color-muted); background: #f8fafc; }}
      th, td {{ padding: 10px 12px; border: 1px solid var(--color-border); vertical-align: top; }}
      ul {{ margin: 0; padding-left: 20px; }}
      .warning-high, .warning-medium, .warning-low {{ padding: 12px 14px; border-radius: 12px; margin-top: 10px; }}
      .warning-high {{ background: #fef2f2; border-left: 4px solid #ef4444; }}
      .warning-medium {{ background: #fffbeb; border-left: 4px solid #f59e0b; }}
      .warning-low {{ background: #f0fdf4; border-left: 4px solid #22c55e; }}
      footer {{ margin-top: 28px; color: var(--color-muted); font-size: 0.85rem; text-align: center; }}
      @media print {{
        body {{ padding: 0; background: #ffffff; font-size: 11pt; }}
        .report {{ max-width: none; }}
        .report-header, .section {{ break-inside: avoid; box-shadow: none; }}
        .metric-grid {{ grid-template-columns: repeat(4, 1fr); }}
      }}
      @media (max-width: 800px) {{
        body {{ padding: 18px; }}
        .metric-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      }}
    </style>
  </head>
  <body>
    <div class="report">
      <header class="report-header">
        <div class="eyebrow">{translate("export.standalone.eyebrow")}</div>
        <h1>{_as_html(request.project_name)}</h1>
        <div class="meta">{_as_html(translate("export.standalone.meta", {"generated_at": generated_at}))}</div>
      </header>
      <div class="section">
        <h2>{translate("export.standalone.key_metrics")}</h2>
        <div class="metric-grid">{metric_cards}</div>
      </div>
      {_render_hypothesis_section(request.hypothesis)}
      {power_curve_svg}
      <div class="section">
        <h2>{translate("export.standalone.assumptions")}</h2>
        <ul>{assumptions_html or f'<li>{translate("export.standalone.none_recorded")}</li>'}</ul>
      </div>
      {_render_design_section(request.design)}
      {warnings_section}
      {sensitivity_svg}
      {ai_section}
      {results_section}
      <footer>{translate("export.standalone.generated_by")}</footer>
    </div>
  </body>
</html>"""


def export_report_to_markdown(report: dict) -> str:
    recommendations = report["recommendations"]
    risks = report["risks"]
    labels = {
        "title": translate("export.markdown.title"),
        "executive_summary": translate("export.markdown.executive_summary"),
        "calculations": translate("export.markdown.calculations"),
        "sample_size_per_variant": translate("export.markdown.sample_size_per_variant"),
        "total_sample_size": translate("export.markdown.total_sample_size"),
        "estimated_duration_days": translate("export.markdown.estimated_duration_days"),
        "assumptions": translate("export.markdown.assumptions"),
        "experiment_design": translate("export.markdown.experiment_design"),
        "randomization_unit": translate("export.markdown.randomization_unit"),
        "traffic_split": translate("export.markdown.traffic_split"),
        "target_audience": translate("export.markdown.target_audience"),
        "inclusion_criteria": translate("export.markdown.inclusion_criteria"),
        "exclusion_criteria": translate("export.markdown.exclusion_criteria"),
        "metrics_plan": translate("export.markdown.metrics_plan"),
        "primary": translate("export.markdown.primary"),
        "secondary": translate("export.markdown.secondary"),
        "guardrail": translate("export.markdown.guardrail"),
        "diagnostic": translate("export.markdown.diagnostic"),
        "risks": translate("export.markdown.risks"),
        "statistical": translate("export.markdown.statistical"),
        "product": translate("export.markdown.product"),
        "technical": translate("export.markdown.technical"),
        "operational": translate("export.markdown.operational"),
        "recommendations": translate("export.markdown.recommendations"),
        "before_launch": translate("export.markdown.before_launch"),
        "during_test": translate("export.markdown.during_test"),
        "after_test": translate("export.markdown.after_test"),
        "open_questions": translate("export.markdown.open_questions"),
    }

    return f"""# {labels["title"]}

## {labels["executive_summary"]}

{report["executive_summary"]}

## {labels["calculations"]}

- {labels["sample_size_per_variant"]}: {report["calculations"]["sample_size_per_variant"]}
- {labels["total_sample_size"]}: {report["calculations"]["total_sample_size"]}
- {labels["estimated_duration_days"]}: {report["calculations"]["estimated_duration_days"]}

### {labels["assumptions"]}

{_list_to_markdown(report["calculations"]["assumptions"])}

## {labels["experiment_design"]}

- {labels["randomization_unit"]}: {report["experiment_design"]["randomization_unit"]}
- {labels["traffic_split"]}: {report["experiment_design"]["traffic_split"]}
- {labels["target_audience"]}: {report["experiment_design"]["target_audience"]}
- {labels["inclusion_criteria"]}: {report["experiment_design"]["inclusion_criteria"]}
- {labels["exclusion_criteria"]}: {report["experiment_design"]["exclusion_criteria"]}

## {labels["metrics_plan"]}

### {labels["primary"]}
{_list_to_markdown(report["metrics_plan"]["primary"])}

### {labels["secondary"]}
{_list_to_markdown(report["metrics_plan"]["secondary"])}

### {labels["guardrail"]}
{_list_to_markdown(report["metrics_plan"]["guardrail"])}

### {labels["diagnostic"]}
{_list_to_markdown(report["metrics_plan"]["diagnostic"])}

## {labels["risks"]}

### {labels["statistical"]}
{_list_to_markdown(risks["statistical"])}

### {labels["product"]}
{_list_to_markdown(risks["product"])}

### {labels["technical"]}
{_list_to_markdown(risks["technical"])}

### {labels["operational"]}
{_list_to_markdown(risks["operational"])}

## {labels["recommendations"]}

### {labels["before_launch"]}
{_list_to_markdown(recommendations["before_launch"])}

### {labels["during_test"]}
{_list_to_markdown(recommendations["during_test"])}

### {labels["after_test"]}
{_list_to_markdown(recommendations["after_test"])}

## {labels["open_questions"]}

{_list_to_markdown(report["open_questions"])}
"""


def export_report_to_html(report: dict) -> str:
    def as_list(items: list[str]) -> str:
        if not items:
            return f"<li>{_as_html(translate('export.html.none'))}</li>"
        return "".join(f"<li>{_as_html(item)}</li>" for item in items)

    recommendations = report["recommendations"]
    risks = report["risks"]
    calculations = report["calculations"]
    design = report["experiment_design"]
    guardrails = report.get("guardrail_metrics", [])
    guardrail_rows = "".join(
        f"<tr><td>{_as_html(item.get('name', ''))}</td><td>{_as_html(item.get('role', ''))}</td><td>{_as_html(item.get('definition', ''))}</td></tr>"
        for item in guardrails
    )
    labels = {
        "title": translate("export.html.title"),
        "meta": translate("export.html.meta"),
        "users_per_variant": translate("export.html.users_per_variant"),
        "total_sample": translate("export.html.total_sample"),
        "estimated_days": translate("export.html.estimated_days"),
        "executive_summary": translate("export.html.executive_summary"),
        "experiment_design": translate("export.html.experiment_design"),
        "randomization_unit": translate("export.html.randomization_unit"),
        "traffic_split": translate("export.html.traffic_split"),
        "target_audience": translate("export.html.target_audience"),
        "inclusion_criteria": translate("export.html.inclusion_criteria"),
        "exclusion_criteria": translate("export.html.exclusion_criteria"),
        "recommended_duration": translate("export.html.recommended_duration"),
        "variants": translate("export.html.variants"),
        "assumptions": translate("export.html.assumptions"),
        "metrics_plan": translate("export.html.metrics_plan"),
        "primary": translate("export.html.primary"),
        "secondary": translate("export.html.secondary"),
        "guardrail": translate("export.html.guardrail"),
        "diagnostic": translate("export.html.diagnostic"),
        "guardrail_metric_definitions": translate("export.html.guardrail_metric_definitions"),
        "name": translate("export.html.name"),
        "role": translate("export.html.role"),
        "definition": translate("export.html.definition"),
        "risks_and_recommendations": translate("export.html.risks_and_recommendations"),
        "statistical": translate("export.html.statistical"),
        "product": translate("export.html.product"),
        "technical": translate("export.html.technical"),
        "operational": translate("export.html.operational"),
        "before_launch": translate("export.html.before_launch"),
        "during_test": translate("export.html.during_test"),
        "after_test": translate("export.html.after_test"),
        "open_questions": translate("export.html.open_questions"),
    }

    return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{labels["title"]}</title>
    <style>
      :root {{
        --color-primary: #0f766e;
        --color-primary-soft: #ccfbf1;
        --color-border: #d7e6df;
        --color-surface: #f8fbf9;
        --color-text: #17312a;
        --color-muted: #4b6b62;
      }}
      * {{ box-sizing: border-box; }}
      body {{ font-family: "Segoe UI", sans-serif; margin: 0; padding: 36px; background: linear-gradient(180deg, #effaf6 0%, #ffffff 180px); color: var(--color-text); }}
      .report {{ max-width: 960px; margin: 0 auto; }}
      .report-header {{ margin-bottom: 24px; padding: 24px; border: 1px solid var(--color-border); border-radius: 24px; background: #ffffff; }}
      h1, h2, h3 {{ color: var(--color-primary); }}
      h1 {{ margin: 0 0 8px; }}
      h2 {{ margin: 0 0 14px; }}
      h3 {{ margin: 16px 0 10px; }}
      p, li, td, th {{ line-height: 1.55; }}
      section {{ margin-bottom: 20px; }}
      .panel {{ border: 1px solid var(--color-border); border-radius: 18px; padding: 20px; background: #ffffff; page-break-inside: avoid; }}
      .summary-grid {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; margin-bottom: 20px; }}
      .summary-card {{ padding: 18px; border-radius: 16px; background: var(--color-surface); text-align: center; }}
      .summary-value {{ font-size: 1.8rem; font-weight: 700; color: var(--color-primary); font-variant-numeric: tabular-nums; }}
      .summary-label {{ margin-top: 6px; color: var(--color-muted); font-size: 0.88rem; }}
      .two-col {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 20px; }}
      table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
      th, td {{ padding: 10px 12px; border: 1px solid var(--color-border); text-align: left; vertical-align: top; }}
      th {{ background: #f2fbf8; color: var(--color-muted); }}
      ul {{ margin: 0; padding-left: 20px; }}
      .meta {{ color: var(--color-muted); }}
      @media print {{
        body {{ padding: 0; background: #ffffff; font-size: 11pt; }}
        .report {{ max-width: none; }}
        .report-header, .panel {{ box-shadow: none; break-inside: avoid; }}
      }}
      @media (max-width: 780px) {{
        body {{ padding: 18px; }}
        .summary-grid, .two-col {{ grid-template-columns: 1fr; }}
      }}
    </style>
  </head>
  <body>
    <div class="report">
      <header class="report-header">
        <h1>{labels["title"]}</h1>
        <div class="meta">{labels["meta"]}</div>
      </header>
      <section class="summary-grid">
        <div class="summary-card">
          <div class="summary-value">{_as_html(calculations["sample_size_per_variant"])}</div>
          <div class="summary-label">{labels["users_per_variant"]}</div>
        </div>
        <div class="summary-card">
          <div class="summary-value">{_as_html(calculations["total_sample_size"])}</div>
          <div class="summary-label">{labels["total_sample"]}</div>
        </div>
        <div class="summary-card">
          <div class="summary-value">{_as_html(calculations["estimated_duration_days"])}</div>
          <div class="summary-label">{labels["estimated_days"]}</div>
        </div>
      </section>
      <section class="panel">
        <h2>{labels["executive_summary"]}</h2>
        <p>{_as_html(report["executive_summary"])}</p>
      </section>
      <section class="panel">
        <h2>{labels["experiment_design"]}</h2>
        <table>
          <tbody>
            <tr><th>{labels["randomization_unit"]}</th><td>{_as_html(design["randomization_unit"])}</td></tr>
            <tr><th>{labels["traffic_split"]}</th><td>{_as_html(design["traffic_split"])}</td></tr>
            <tr><th>{labels["target_audience"]}</th><td>{_as_html(design["target_audience"])}</td></tr>
            <tr><th>{labels["inclusion_criteria"]}</th><td>{_as_html(design["inclusion_criteria"])}</td></tr>
            <tr><th>{labels["exclusion_criteria"]}</th><td>{_as_html(design["exclusion_criteria"])}</td></tr>
            <tr><th>{labels["recommended_duration"]}</th><td>{_as_html(design["recommended_duration_days"])}</td></tr>
          </tbody>
        </table>
        <h3>{labels["variants"]}</h3>
        <ul>{as_list([f'{item["name"]}: {item["description"]}' for item in design["variants"]])}</ul>
        <h3>{labels["assumptions"]}</h3>
        <ul>{as_list(calculations["assumptions"])}</ul>
      </section>
      <section class="panel">
        <h2>{labels["metrics_plan"]}</h2>
        <div class="two-col">
          <div>
            <h3>{labels["primary"]}</h3>
            <ul>{as_list(report["metrics_plan"]["primary"])}</ul>
            <h3>{labels["secondary"]}</h3>
            <ul>{as_list(report["metrics_plan"]["secondary"])}</ul>
          </div>
          <div>
            <h3>{labels["guardrail"]}</h3>
            <ul>{as_list(report["metrics_plan"]["guardrail"])}</ul>
            <h3>{labels["diagnostic"]}</h3>
            <ul>{as_list(report["metrics_plan"]["diagnostic"])}</ul>
          </div>
        </div>
        {f'<h3>{labels["guardrail_metric_definitions"]}</h3><table><thead><tr><th>{labels["name"]}</th><th>{labels["role"]}</th><th>{labels["definition"]}</th></tr></thead><tbody>{guardrail_rows}</tbody></table>' if guardrail_rows else ''}
      </section>
      <section class="panel">
        <h2>{labels["risks_and_recommendations"]}</h2>
        <div class="two-col">
          <div>
            <h3>{labels["statistical"]}</h3>
            <ul>{as_list(risks["statistical"])}</ul>
            <h3>{labels["product"]}</h3>
            <ul>{as_list(risks["product"])}</ul>
            <h3>{labels["technical"]}</h3>
            <ul>{as_list(risks["technical"])}</ul>
            <h3>{labels["operational"]}</h3>
            <ul>{as_list(risks["operational"])}</ul>
          </div>
          <div>
            <h3>{labels["before_launch"]}</h3>
            <ul>{as_list(recommendations["before_launch"])}</ul>
            <h3>{labels["during_test"]}</h3>
            <ul>{as_list(recommendations["during_test"])}</ul>
            <h3>{labels["after_test"]}</h3>
            <ul>{as_list(recommendations["after_test"])}</ul>
          </div>
        </div>
      </section>
      <section class="panel">
        <h2>{labels["open_questions"]}</h2>
        <ul>{as_list(report["open_questions"])}</ul>
      </section>
    </div>
  </body>
</html>
"""


def _raw_input_rows(payload: dict[str, Any], *, prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    for key in sorted(payload.keys()):
        value = payload[key]
        next_prefix = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            rows.extend(_raw_input_rows(value, prefix=next_prefix))
            continue
        rows.append((next_prefix, value))
    return rows


def _build_report_export_sections(project: dict[str, Any], analysis: dict[str, Any]) -> dict[str, Any]:
    payload = project.get("payload", {})
    calculations = analysis.get("calculations", {})
    report = analysis.get("report", {})
    experiment_design = report.get("experiment_design", {})
    setup = payload.get("setup", {})
    variants = experiment_design.get("variants", [])
    traffic_split = experiment_design.get("traffic_split", setup.get("traffic_split", []))
    results = calculations.get("results", {})
    sensitivity_rows = _build_sensitivity_rows(payload, calculations)
    sample_rows = [
        [
            str(variant.get("name", f"Variant {index + 1}")),
            results.get("sample_size_per_variant"),
            results.get("effective_daily_traffic"),
            results.get("estimated_duration_days"),
            traffic_split[index] if index < len(traffic_split) else None,
        ]
        for index, variant in enumerate(variants)
    ]
    guardrail_rows = [
        [
            item.get("name"),
            item.get("metric_type"),
            item.get("detectable_mde_pp") or item.get("detectable_mde_absolute"),
            item.get("note"),
        ]
        for item in report.get("guardrail_metrics", [])
    ]
    return {
        "sample_rows": sample_rows,
        "sensitivity_rows": sensitivity_rows,
        "guardrail_rows": guardrail_rows,
        "raw_input_rows": _raw_input_rows(payload),
    }


def export_project_report_to_csv(project: dict[str, Any], analysis: dict[str, Any]) -> str:
    sections = _build_report_export_sections(project, analysis)
    buffer = StringIO()
    writer = csv.writer(buffer, lineterminator="\n")

    writer.writerow([f'# {translate("export.project.sample_results_header")}'])
    writer.writerow([
        translate("export.project.variant"),
        translate("export.project.required_n"),
        translate("export.project.daily_traffic"),
        translate("export.project.duration_days"),
        translate("export.project.traffic_split_pct"),
    ])
    writer.writerows(sections["sample_rows"])
    writer.writerow([])

    writer.writerow([f'# {translate("export.project.sensitivity_header")}'])
    writer.writerow([
        translate("export.project.effect_size"),
        translate("export.project.required_n_80"),
        translate("export.project.duration_80"),
        translate("export.project.required_n_90"),
        translate("export.project.duration_90"),
    ])
    for row in sections["sensitivity_rows"]:
        writer.writerow([
            row.get("effect_size"),
            row.get("n_80"),
            row.get("days_80"),
            row.get("n_90"),
            row.get("days_90"),
        ])
    writer.writerow([])

    writer.writerow([f'# {translate("export.project.guardrails_header")}'])
    writer.writerow([
        translate("export.project.metric_name"),
        translate("export.project.metric_type"),
        translate("export.project.detectable_mde"),
        translate("export.project.note"),
    ])
    writer.writerows(sections["guardrail_rows"])

    return buffer.getvalue()


def export_project_report_to_xlsx(project: dict[str, Any], analysis: dict[str, Any]) -> bytes:
    sections = _build_report_export_sections(project, analysis)
    payload = project.get("payload", {})
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = translate("export.project.summary_sheet")
    summary_sheet.append([translate("export.project.project_name"), project.get("project_name")])
    summary_sheet.append([translate("export.project.created_at"), project.get("created_at")])
    summary_sheet.append([])
    summary_sheet.append([
        translate("export.project.variant"),
        translate("export.project.required_n"),
        translate("export.project.daily_traffic"),
        translate("export.project.duration_days"),
        translate("export.project.traffic_split_pct"),
    ])
    for row in sections["sample_rows"]:
        summary_sheet.append(row)

    sensitivity_sheet = workbook.create_sheet(translate("export.project.sensitivity_sheet"))
    sensitivity_sheet.append([
        translate("export.project.effect_size"),
        translate("export.project.required_n_80"),
        translate("export.project.duration_80"),
        translate("export.project.required_n_90"),
        translate("export.project.duration_90"),
    ])
    for row in sections["sensitivity_rows"]:
        sensitivity_sheet.append([
            row.get("effect_size"),
            row.get("n_80"),
            row.get("days_80"),
            row.get("n_90"),
            row.get("days_90"),
        ])

    guardrails_sheet = workbook.create_sheet(translate("export.project.guardrails_sheet"))
    guardrails_sheet.append([
        translate("export.project.metric_name"),
        translate("export.project.metric_type"),
        translate("export.project.detectable_mde"),
        translate("export.project.note"),
    ])
    for row in sections["guardrail_rows"]:
        guardrails_sheet.append(row)

    raw_inputs_sheet = workbook.create_sheet(translate("export.project.raw_inputs_sheet"))
    raw_inputs_sheet.append([translate("export.project.field"), translate("export.project.value")])
    for field, value in sections["raw_input_rows"]:
        if isinstance(value, list):
            raw_inputs_sheet.append([field, json.dumps(value)])
        else:
            raw_inputs_sheet.append([field, value])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
